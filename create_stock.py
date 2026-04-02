import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════
#  FEUILLE CATALOGUE
# ════════════════════════════════════════════════
ws = wb.active
ws.title = "Catalogue"

# Couleurs
HEADER_COLOR   = "2C3E50"
CAT_COLORS = {
    "BOISSONS":              "2980B9",
    "BOISSONS CHAUDES":      "8E44AD",
    "LAIT/SUCRE":            "16A085",
    "POTAGES":               "D35400",
    "ENCAS SALES":           "C0392B",
    "BISCUITS/CONFISERIES":  "E67E22",
    "GOURMALLIANCE":         "27AE60",
    "DIVERS":                "7F8C8D",
}

# En-têtes
headers = ["REF", "DESIGNATION", "CATEGORIE", "STOCK ACTUEL", "STOCK MINI", "STOCK MAXI", "DERNIER ARRIVAGE"]
ws.append(headers)

header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
for col in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=col)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Données ──
produits = [
    # (REF, DESIGNATION, CATEGORIE)
    # BOISSONS
    ("BOISSONS", None, None),
    ("CAPRISUN",       "CAPRISUN MULTIVIT POCHE 20CL *40",               "BOISSONS"),
    ("50COCA",         "COCA COLA 50CL *24",                             "BOISSONS"),
    ("33COCHE1",       "COCA COLA CHERRY SLIM CAN 33CL *24",             "BOISSONS"),
    ("33COCASLIM",     "COCA COLA SLIM CAN 33CL *24",                    "BOISSONS"),
    ("50CRISTACV",     "CRISTALINE CITRON PET 50CL *24",                 "BOISSONS"),
    ("50CRISTAP",      "CRISTALINE EAU PLATE 50CL *24",                  "BOISSONS"),
    ("50CRISTAFR",     "CRISTALINE FRUITS ROUGES PET 50CL *24",          "BOISSONS"),
    ("50CRISTA",       "CRISTALINE GAZEIFIE 50CL *24",                   "BOISSONS"),
    ("33FANORA1",      "FANTA ORANGE SLIM 33CL *24",                     "BOISSONS"),
    ("FUZTPI",         "FUZETEA PECHE INTENSE 40CL *12",                 "BOISSONS"),
    ("33FUZETEAUP1",   "FUZETEA PECHE INTENSE SLIM 33CL *24",            "BOISSONS"),
    ("33ICEPEC",       "ICE TEA LIPTON PECHE 33CL SLIM *24",             "BOISSONS"),
    ("33OAPOCA1",      "OASIS POM CASSIS FRAMB SLIM 33CL *24",           "BOISSONS"),
    ("50OASISTROPICAL","OASIS TROPICAL 50CL *12",                        "BOISSONS"),
    ("33OASTROSLIM",   "OASIS TROPICAL SLIM 33CL *24",                   "BOISSONS"),
    ("33ORAJAU",       "ORANGINA JAUNE SLIM 33CL *24",                   "BOISSONS"),
    ("ORANGINAJ",      "ORANGINA JAUNE 50CL *12",                        "BOISSONS"),
    ("50PEPSI",        "PEPSI ZERO 50CL *12",                            "BOISSONS"),
    ("PEPMAXSLIM",     "PEPSI MAX SLIM 33CL *24",                        "BOISSONS"),
    ("33PERCV",        "PERRIER CITRVER EAU GAZ SLIM 33CL *24",          "BOISSONS"),
    ("33PERR1",        "PERRIER NATURE EAU GAZ SLIM 33CL *24",           "BOISSONS"),
    ("25REDBULL",      "REDBULL SLIM 25CL *24",                          "BOISSONS"),
    ("1LSANPEL",       "SAN PELLIGRINO 1L *6",                           "BOISSONS"),
    ("50SANPEL",       "SAN PELLIGRINO 50CL *24",                        "BOISSONS"),
    ("33SEVENUPSLIM",  "SEVEN UP SLIM 33CL *24",                         "BOISSONS"),
    ("33SCHAGR1",      "SCHWEPPES AGRUM SLIM 33CL *24",                  "BOISSONS"),
    ("33SPRIVAL",      "SPRING VALLEY JUS TROPICAL SLIM 33CL *24",       "BOISSONS"),
    ("33SPRINGVO1",    "SPRING VALLEY ORANGE SLIM 33CL *24",             "BOISSONS"),
    ("33SPRINGVP",     "SPRING VALLEY POMME SLIM 33CL *24",              "BOISSONS"),
    ("33SPRITE",       "SPRITE SLIM 33CL *24",                           "BOISSONS"),
    ("50VOLVIC",       "VOLVIC 50CL *24",                                "BOISSONS"),

    # BOISSONS CHAUDES
    ("BOISSONS CHAUDES", None, None),
    ("CAN07G",   "CAFE EXTRA EXPRESSO N°07 GRAINS",                      "BOISSONS CHAUDES"),
    ("CALYMAX",  "CAFE LYOPHILISE MAXWELL SP.FILTRE 500G",               "BOISSONS CHAUDES"),
    ("CALYDECA", "CAFE LYOPHILISE RISTRETTO DECA POCHE",                 "BOISSONS CHAUDES"),
    ("CANO3G",   "CAFE SUPERIEUR EXPRESSO N°03 GRAINS",                  "BOISSONS CHAUDES"),
    ("CHOC",     "CHOCOLAT VH1 VAN HOUTEN D.A. *10",                     "BOISSONS CHAUDES"),
    ("THECIT",   "THE CITRON DA 1KG *10",                                "BOISSONS CHAUDES"),
    ("THEFRU",   "THE FRUITS ROUGES DA 1KG *10",                         "BOISSONS CHAUDES"),
    ("THEMEN",   "THE MENTHE DA 1KG *10",                                "BOISSONS CHAUDES"),
    ("CAFCAR",   "VENDING CAFE CARAMEL DA 1KG *10",                      "BOISSONS CHAUDES"),
    ("CAFNOI",   "VENDING CAPUCCINO NOISETTE 1KG *10",                   "BOISSONS CHAUDES"),
    ("CAFVAN",   "VENDING CAPUCCINO VANILLE 1KG *10",                    "BOISSONS CHAUDES"),

    # LAIT/SUCRE
    ("LAIT/SUCRE", None, None),
    ("LAITNS",  "LAIT DA NON SUCRE 500G *10",  "LAIT/SUCRE"),
    ("SUCSEM",  "SUCRE SEMOULE D.A. 2KG *6",   "LAIT/SUCRE"),

    # POTAGES
    ("POTAGES", None, None),
    ("DNLPOTTEXM", "DRINK A LIKE POTAGE TEX MEX 1KG *5",         "POTAGES"),
    ("DNLPOTPOI",  "DRINK N LIKE POTAGE POIREAU PDT 1KG *5",     "POTAGES"),
    ("DNLPOTPROV", "DRINK N LIKE POTAGE PROVENCAL 1KG *5",       "POTAGES"),
    ("DNLPOTTOM",  "DRINK N LIKE POTAGE TOMATE 1KG *5",          "POTAGES"),

    # ENCAS SALES
    ("ENCAS SALES", None, None),
    ("COCRAF",   "CRACKERS SAVEUR FROMAGE ET POIVRE 40G *50",    "ENCAS SALES"),
    ("COCRAR",   "CRAKERS ROMARIN PATATE *50",                   "ENCAS SALES"),
    ("COCRAT",   "CRAKERS TOMATE ORIGAN *50",                    "ENCAS SALES"),
    ("COPATPOU", "PATES POULET *12",                             "ENCAS SALES"),
    ("COSALNA",  "SALADE NAPOLITAINE THON *12",                  "ENCAS SALES"),
    ("COSALPI",  "SALADE PIEMONTAISE THON *12",                  "ENCAS SALES"),
    ("COSALRIZ", "SALADE RIZ THON 220GR *7",                     "ENCAS SALES"),
    ("COSALVG",  "SALADE VEGETARIENNE 220G *7",                  "ENCAS SALES"),
    ("SALPRO",   "SALADE PROVENCALE AU POULET *7",               "ENCAS SALES"),

    # BISCUITS/CONFISERIES
    ("BISCUITS/CONFISERIES", None, None),
    ("COBALRAINOI", "BALISTO CEREALES RAISIN NOISETTE 37G *20",           "BISCUITS/CONFISERIES"),
    ("COBALISCER",  "BALISTO MIEL AMANDE 37G *20",                        "BISCUITS/CONFISERIES"),
    ("COBF2",       "BARRE DE FRUITS POMME ORANGE CAROTTE 45GR *20",      "BISCUITS/CONFISERIES"),
    ("COBEL",       "BELVITA PETIT DEJEUNER 50G *30",                     "BISCUITS/CONFISERIES"),
    ("COBOULAI",    "BOUNTY CHOCO 57G *24",                               "BISCUITS/CONFISERIES"),
    ("SACHIPSFLO",  "CHIPS SEL FLODOR 30G *40",                           "BISCUITS/CONFISERIES"),
    ("COTEDORLN",   "COTE D'OR LAIT NOISETTE BARRE CHOCOLATEE *32",       "BISCUITS/CONFISERIES"),
    ("COCRUNCH",    "CRUNCH 30G *30",                                     "BISCUITS/CONFISERIES"),
    ("CURLYCA",     "CURLY CACAHUETES 60G *30",                           "BISCUITS/CONFISERIES"),
    ("GALAMA",      "GALETTE MOELLEUSE SAVEUR AMANDE *24",                "BISCUITS/CONFISERIES"),
    ("COGALC",      "GALETTE RIZ CHOCO *20",                              "BISCUITS/CONFISERIES"),
    ("GAULIE",      "GAUFRE LIEGEOISE 90G *30",                           "BISCUITS/CONFISERIES"),
    ("GERCRAAM",    "GERBLE BIO CRANBERRY AMANDE *18",                    "BISCUITS/CONFISERIES"),
    ("GERBLEBIOGC", "GERBLE BISCUIT BIO GINGEMBRE CITRON *18",            "BISCUITS/CONFISERIES"),
    ("COGEBI",      "GERBLE BISCUIT LAIT CHOCOLAT 46G *18",               "BISCUITS/CONFISERIES"),
    ("COGESE",      "GERBLE BISCUIT SESAME 46G *18",                      "BISCUITS/CONFISERIES"),
    ("CROGER",      "GERBLE CROQUANT MIEL SESAME 27G *18",                "BISCUITS/CONFISERIES"),
    ("GERGALCHO",   "GERBLE GALETTE MAIS CHOCO LAIT *56",                 "BISCUITS/CONFISERIES"),
    ("GERCHOCLAIT", "GERBLE GALETTE MAIS CHOCOLAT LAIT *28",              "BISCUITS/CONFISERIES"),
    ("GALGERN",     "GERBLE GALETTE RIZ CHOCO NOIR 33G *28",              "BISCUITS/CONFISERIES"),
    ("HARDELIR",    "HARIBO SACHET 120G *30",                             "BISCUITS/CONFISERIES"),
    ("BARSPEKC",    "KELLOGS SPECIAL K CHOCOLAT 21.5G *30",               "BISCUITS/CONFISERIES"),
    ("BARSPEK",     "KELLOGS SPECIAL K FRUIT ROUGE 21.5G *30",            "BISCUITS/CONFISERIES"),
    ("COKINBUEW",   "KINDER BUENO WHITE *30",                             "BISCUITS/CONFISERIES"),
    ("COKINBUE",    "KINDER BUENO 43G *30",                               "BISCUITS/CONFISERIES"),
    ("COKITKATC",   "KIT KAT CHUNKY 40G *24",                             "BISCUITS/CONFISERIES"),
    ("COLON",       "KIT KAT FINGER 41.5G *36",                           "BISCUITS/CONFISERIES"),
    ("COLIONC",     "LION 42G *24",                                       "BISCUITS/CONFISERIES"),
    ("COLIONC15",   "LION BARRE CHOCO +15% *24",                          "BISCUITS/CONFISERIES"),
    ("COMM50",      "M&M'S CHOCO PEANUT 45G *36",                         "BISCUITS/CONFISERIES"),
    ("MADNAT",      "MADELEINE NATURE 84G *24",                           "BISCUITS/CONFISERIES"),
    ("COMAL",       "MALTESERS 37G *25",                                  "BISCUITS/CONFISERIES"),
    ("COMARS",      "MARS CHOCO 51G *40",                                 "BISCUITS/CONFISERIES"),
    ("COMAGA",      "MAXI GALETTE 100G *24",                              "BISCUITS/CONFISERIES"),
    ("MIKADOLAIT",  "MIKADO POCKET LAIT *24",                             "BISCUITS/CONFISERIES"),
    ("MINISAVANE",  "MINI SAVANE CHOCOLAT X2 60G *36",                    "BISCUITS/CONFISERIES"),
    ("GOMUNATC",    "MUFFIN NATURE & PEPITES CHOCOLAT *40",               "BISCUITS/CONFISERIES"),
    ("NAPO60GX2",   "NAPOLITAIN CLASSIC X2 60G *24",                      "BISCUITS/CONFISERIES"),
    ("NATVALCHO",   "NATURE VALLEY&DARK NUT CHO 30G *18",                 "BISCUITS/CONFISERIES"),
    ("CONUTS",      "NUTELLA NOISETTE CHOCO 42G *24",                     "BISCUITS/CONFISERIES"),
    ("COOREO",      "OREO 66G *20",                                       "BISCUITS/CONFISERIES"),
    ("PETITECO",    "PETIT ECOLIER CHOCO LAIT *23",                       "BISCUITS/CONFISERIES"),
    ("POMPOTEPOM",  "POMPOTE ANA GRENADE SSA *48",                        "BISCUITS/CONFISERIES"),
    ("COPRINCE",    "PRINCE CHOCOLAT 80G *20",                            "BISCUITS/CONFISERIES"),
    ("COSNIC",      "SNICKERS 50G *40",                                   "BISCUITS/CONFISERIES"),
    ("COTWIX",      "TWIX 50G *32",                                       "BISCUITS/CONFISERIES"),

    # GOURMALLIANCE
    ("GOURMALLIANCE", None, None),
    ("COCECH",    "BARRE DE CEREALES AU CHOCOLAT *20",                    "GOURMALLIANCE"),
    ("COCENO",    "BARRE DE CEREALES AUX NOISETTES ENKA *20",             "GOURMALLIANCE"),
    ("COCECF",    "BARRE DE CEREALES CRANB ET FRAM *20",                  "GOURMALLIANCE"),
    ("COGOUP",    "GOURDE MIXE DE POMMES BIO 90GR *55",                   "GOURMALLIANCE"),
    ("COCOOKM62", "GRAND COOKIE MOELLEUX BEURRE CACAH MILLET & CHIA *26", "GOURMALLIANCE"),
    ("COCOAM",    "GRAND COOKIE MOELLEUX BROWNIE CHOCO *26",              "GOURMALLIANCE"),
    ("COCOOK",    "GRAND COOKIE MOELLEUX COCO AMANDES *26",               "GOURMALLIANCE"),
    ("COGAULIE",  "GRANDE GAUFFRE LIEGOISE SUCREE MI CHOCO *24",          "GOURMALLIANCE"),
    ("MIXTON",    "MIX TONUS BIO 35G *12",                                "GOURMALLIANCE"),
    ("GOUPP",     "GOURDE MIXE *10",                                      "GOURMALLIANCE"),
    ("PICKUP",    "PICKUP *24",                                           "GOURMALLIANCE"),
    ("MINCOOK",   "MINI COOKIE 80G *75",                                  "GOURMALLIANCE"),
    ("PITAPIZ",   "MINI PITA 30G *50",                                    "GOURMALLIANCE"),

    # DIVERS
    ("DIVERS", None, None),
    ("GOB151924",  "GOBELET PLASTIQUE DA BLANC 15CL 30R *100",            "DIVERS"),
    ("CHLORE",     "PASTILLE JAVEL-500G-150 COMPRIMES",                   "DIVERS"),
    ("WH20006021", "PRODUIT LESSIVIEL POLYVALENT F300 5L",                "DIVERS"),
    ("SACPOU50L",  "SAC POUBELLE 50L",                                    "DIVERS"),
    ("SPATUL",     "SPATULES BOIS 105MM 60R *50",                         "DIVERS"),
    ("HYESSTOU",   "PAPIER ESSUIE MAIN",                                  "DIVERS"),
]

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_num = 2
for item in produits:
    ref, desig, cat = item

    # Ligne de catégorie
    if desig is None:
        color = CAT_COLORS.get(ref, "555555")
        ws.cell(row=row_num, column=1).value = ref
        ws.merge_cells(f"A{row_num}:G{row_num}")
        for col in range(1, 8):
            c = ws.cell(row=row_num, column=col)
            c.fill      = PatternFill("solid", fgColor=color)
            c.font      = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
        row_num += 1
        continue

    # Ligne produit
    ws.cell(row=row_num, column=1).value = ref
    ws.cell(row=row_num, column=2).value = desig
    ws.cell(row=row_num, column=3).value = cat
    ws.cell(row=row_num, column=4).value = 0   # stock actuel
    ws.cell(row=row_num, column=5).value = 0   # stock mini
    ws.cell(row=row_num, column=6).value = 0   # stock maxi
    ws.cell(row=row_num, column=7).value = ""  # dernier arrivage

    bg = "FFFFFF" if row_num % 2 == 0 else "F8F9FA"
    for col in range(1, 8):
        c = ws.cell(row=row_num, column=col)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center")
        c.border    = border

    row_num += 1

# Largeurs colonnes
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 20

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# ════════════════════════════════════════════════
#  FEUILLE HISTORIQUE
# ════════════════════════════════════════════════
ws_h = wb.create_sheet("Historique")
hist_headers = ["ID Commande", "Date", "Heure", "Destination",
                "Code Produit", "Désignation", "Qté Prélevée", "Stock Après"]
ws_h.append(hist_headers)

for col in range(1, len(hist_headers)+1):
    c = ws_h.cell(row=1, column=col)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
    c.alignment = Alignment(horizontal="center", vertical="center")

ws_h.column_dimensions['A'].width = 28
ws_h.column_dimensions['B'].width = 14
ws_h.column_dimensions['C'].width = 10
ws_h.column_dimensions['D'].width = 12
ws_h.column_dimensions['E'].width = 18
ws_h.column_dimensions['F'].width = 45
ws_h.column_dimensions['G'].width = 14
ws_h.column_dimensions['H'].width = 14
ws_h.freeze_panes = "A2"

wb.save("stock.xlsx")
print("✅ stock.xlsx créé avec succès !")
