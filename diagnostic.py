# diagnostic.py
import openpyxl

EXCEL_FILE = 'stock.xlsx'

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb["Catalogue"]

print(f"Nombre de lignes : {ws.max_row}")
print()

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    print(f"Ligne {i+2} : {row}")

wb.close()
