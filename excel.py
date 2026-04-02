import openpyxl
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime
from config import EXCEL_FILE

# ════════════════════════════════════════════════
#  UTILITAIRES
# ════════════════════════════════════════════════

def normalise_code(code):
    c = str(code).strip()
    if 'E+' in c or 'e+' in c:
        try:
            c = str(int(float(c)))
        except:
            pass
    return c

def get_workbook():
    return openpyxl.load_workbook(EXCEL_FILE)

# ════════════════════════════════════════════════
#  CATALOGUE
# ════════════════════════════════════════════════

def load_catalogue():
    wb = get_workbook()
    ws = wb["Catalogue"]
    catalogue = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ignorer lignes vides
        if not row[0]:
            continue

        code_raw = str(row[0]).strip()  # ← Colonne A = REF

        # Ignorer en-têtes et séparateurs de catégorie
        if code_raw.lower() in ['ref', 'code', 'n°', '']:
            continue
        # Ignorer les lignes de catégorie (ex: "BOISSONS", "LAIT/SUCRE"...)
        if row[1] is None:
            continue

        code = normalise_code(code_raw)

        try:
            nom         = str(row[1]).strip() if row[1] else ''   # B: DESIGNATION
            categorie   = str(row[2]).strip() if row[2] else ''   # C: CATEGORIE
            stock       = int(row[3]) if row[3] not in (None, '') else 0  # D: STOCK ACTUEL
            stock_mini  = int(row[4]) if row[4] not in (None, '') else 0  # E: STOCK MINI
            stock_maxi  = int(row[5]) if row[5] not in (None, '') else 0  # F: STOCK MAXI
            dernier_arrivage = str(row[6]) if row[6] else ''              # G: DERNIER ARRIVAGE
        except Exception as e:
            print(f"⚠ Catalogue — ligne ignorée ({row[1]}) : {e}")
            continue

        catalogue[code] = {
            'code':             code,
            'nom':              nom,
            'categorie':        categorie,
            'stock':            stock,
            'stock_mini':       stock_mini,
            'stock_maxi':       stock_maxi,
            'dernier_arrivage': dernier_arrivage,
        }

    wb.close()
    return catalogue



def update_catalogue_row(ws, code, stock_apres, qty):
    """
    Col A=REF, B=Désignation, C=Catégorie, D=Stock Actuel, E=Stock Mini, F=Stock Maxi
    Met à jour le stock dans la colonne D.
    """
    for row in ws.iter_rows(min_row=2):
        if not row[0].value:
            continue
        cell_code = normalise_code(str(row[0].value))
        if cell_code == code:
            row[3].value = stock_apres  # colonne D = Stock Actuel
            return True
    return False


# ════════════════════════════════════════════════
#  HISTORIQUE
# ════════════════════════════════════════════════

def append_historique_row(ws, cmd_id, dest, code, nom, qty, stock_apres):
    """
    Ajoute une ligne dans la feuille Historique.
    Colonnes : CMD_ID | Date | Heure | Destination | Code | Nom | Quantité | Stock après
    """
    now = datetime.now()
    ws.append([
        cmd_id,
        now.strftime('%Y-%m-%d'),
        now.strftime('%H:%M:%S'),
        dest,
        code,
        nom,
        qty,
        stock_apres,
    ])

    # Alternance couleur de ligne
    last = ws.max_row
    fill = PatternFill("solid", fgColor="F2F2F2" if last % 2 == 0 else "FFFFFF")
    for cell in ws[last]:
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')


def save_scan_to_excel(cmd_id, destination, code, nom, quantite, stock_apres):
    """
    Sauvegarde un scan :
    - Met à jour le stock dans Catalogue
    - Ajoute la ligne dans Historique
    """
    wb = get_workbook()
    update_catalogue_row(wb["Catalogue"], code, stock_apres, quantite)
    append_historique_row(wb["Historique"], cmd_id, destination, code, nom, quantite, stock_apres)
    wb.save(EXCEL_FILE)
    wb.close()


def load_historique():
    """
    Lit la feuille Historique et regroupe les lignes par commande.
    Retourne une liste de commandes triées par date/heure décroissante.
    """
    wb = get_workbook()
    ws = wb["Historique"]
    commandes = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        cmd_id = str(row[0]).strip()

        # Ignorer la ligne d'en-tête si elle existe
        if cmd_id.lower() in ['id commande', 'cmd_id', 'id']:
            continue

        try:
            date        = str(row[1]) if row[1] else ''
            heure       = str(row[2]) if row[2] else ''
            dest        = str(row[3]) if row[3] else ''
            code        = str(row[4]) if row[4] else ''
            nom         = str(row[5]) if row[5] else ''
            qty         = int(row[6]) if row[6] not in (None, '') else 0
            stock_apres = int(row[7]) if row[7] not in (None, '') else 0
        except Exception as e:
            print(f"⚠ Historique — ligne ignorée ({cmd_id}) : {e}")
            continue

        if cmd_id not in commandes:
            commandes[cmd_id] = {
                'id':          cmd_id,
                'date':        date,
                'heure':       heure,
                'destination': dest,
                'produits':    [],
                'nb_articles': 0,
            }

        commandes[cmd_id]['produits'].append({
            'code':        code,
            'nom':         nom,
            'quantite':    qty,
            'stock_apres': stock_apres,
        })
        commandes[cmd_id]['nb_articles'] += qty

    wb.close()

    # Trier du plus récent au plus ancien
    return sorted(commandes.values(), key=lambda c: (c['date'], c['heure']), reverse=True)


def update_stock_in_excel(code, new_stock):
    wb = get_workbook()
    ws = wb["Catalogue"]
    for row in ws.iter_rows(min_row=2):
        if not row[0].value:
            continue
        if normalise_code(str(row[0].value)) == code:
            row[3].value = new_stock  # colonne D = Stock Actuel ✅
            break
    wb.save(EXCEL_FILE)
    wb.close()

