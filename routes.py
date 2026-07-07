from flask import Blueprint, render_template, request, jsonify, send_file
from datetime import datetime
import urllib.request
import json
from models import db, Produit, Historique, CommandeDA, CommandeFournisseur

from config import (
    DESTINATIONS, DEST_COLORS_HEX,
    MAIL_EXPEDITEUR, MAIL_DESTINATAIRE, BREVO_API_KEY
)
from pdf import generate_commande_pdf, generate_journalier_pdf, generate_commande_da_pdf, generate_commande_fournisseur_pdf

bp = Blueprint('main', __name__)


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def calculer_stock_provisoire(code_produit):
    """Retourne le stock réel - les quantités réservées par les DA en attente."""
    from sqlalchemy import func
    total_da = db.session.query(func.coalesce(func.sum(CommandeDA.quantite), 0))\
        .filter(CommandeDA.code == code_produit, CommandeDA.statut == 'en_attente').scalar()
    produit = Produit.query.get(code_produit)
    if not produit:
        return 0
    return max(0, produit.stock - total_da)


def normalise_code(code):
    c = str(code).strip()
    if 'E+' in c or 'e+' in c:
        try:
            c = str(int(float(c)))
        except:
            pass
    return c


def nouvelle_commande_id():
    return datetime.now().strftime('CMD-%Y%m%d-%H%M%S')


def load_historique():
    rows = Historique.query.order_by(
        Historique.date.desc(), Historique.heure.desc()
    ).all()
    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':          r.cmd_id,
                'date':        r.date,
                'heure':       r.heure,
                'destination': r.destination,
                'produits':    [],
                'nb_articles': 0,
            }
        commandes[r.cmd_id]['produits'].append({
            'code':        r.code,
            'nom':         r.nom,
            'quantite':    r.quantite,
            'stock_apres': r.stock_apres,
        })
        commandes[r.cmd_id]['nb_articles'] += r.quantite

    return sorted(
        commandes.values(),
        key=lambda c: (c['date'], c['heure']),
        reverse=True
    )


def envoyer_mail_commande(commande, action="nouvelle"):
    if not MAIL_EXPEDITEUR or not MAIL_DESTINATAIRE or not MAIL_DESTINATAIRE1 or not BREVO_API_KEY:
        print("⚠️  Mail non configuré (BREVO_API_KEY manquante), envoi ignoré")
        return

    try:
        action_label = {
            'nouvelle': 'NOUVELLE COMMANDE',
            'modifiée': 'COMMANDE MODIFIÉE',
        }.get(action, action.upper())

        sujet = f"[{action_label}] {commande['id']} — {commande['destination']}"

        # Corps HTML simple
        lignes = [f"<b>Commande :</b> {commande['id']}<br>",
                  f"<b>Destination :</b> {commande['destination']}<br>",
                  f"<b>Date :</b> {commande['date']} à {commande['heure']}<br>",
                  f"<b>Action :</b> {action}<br><br>",
                  "<b>Produits :</b><br><ul>"]
        for p in commande['produits']:
            lignes.append(f"<li>{p.get('nom', '')} ({p.get('code', '')}) x{p.get('quantite', 0)}</li>")
        lignes.append("</ul>")
        corps_html = "".join(lignes)

        # Corps texte brut
        corps_txt_lignes = [
            f"Commande   : {commande['id']}",
            f"Destination: {commande['destination']}",
            f"Date       : {commande['date']} à {commande['heure']}",
            f"Action     : {action}",
            "",
            "Produits :",
        ]
        for p in commande['produits']:
            corps_txt_lignes.append(f"  - {p.get('nom', '')} ({p.get('code', '')}) x{p.get('quantite', 0)}")
        corps_txt = "\n".join(corps_txt_lignes)

        # Vérification débogage
        print(f"📧 Envoi mail via Brevo API...")
        print(f"   Expéditeur: {MAIL_EXPEDITEUR}")
        print(f"   Destinataire: {MAIL_DESTINATAIRE}")
        print(f"   Destinataire1: {MAIL_DESTINATAIRE1}")
        print(f"   Clé API (début): {BREVO_API_KEY[:10]}...")
        print(f"   Sujet: {sujet}")

        # Appel API Brevo
        url = "https://api.brevo.com/v3/smtp/email"
        headers = {
            "Content-Type": "application/json",
            "api-key": BREVO_API_KEY,
        }

        payload = {
            "sender": {"email": MAIL_EXPEDITEUR, "name": "Stock Scan"},
            "to": [{"email": MAIL_DESTINATAIRE},{"email": MAIL_DESTINATAIRE1}],
            "subject": sujet,
            "htmlContent": corps_html,
            "textContent": corps_txt,
        }

        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers=headers, method='POST')
        resp = urllib.request.urlopen(req, timeout=15)
        status = resp.getcode()
        body = resp.read().decode('utf-8')

        if status == 201:
            print(f"✅ Mail envoyé via Brevo pour {commande['id']} ({action})")
        else:
            print(f"⚠️ Brevo a répondu {status}: {body}")

    except urllib.error.HTTPError as e:
        print(f"❌ Erreur HTTP {e.code}: {e.read().decode('utf-8')}")
    except Exception as e:
        print(f"❌ Erreur envoi mail Brevo : {e}")


# ══════════════════════════════════════════════════════════════
#  PAGES
# ══════════════════════════════════════════════════════════════

@bp.route('/')
def index():
    return render_template(
        'index.html',
        destinations=DESTINATIONS,
        dest_colors=DEST_COLORS_HEX
    )


@bp.route('/commandes_da')
def commandes_da_page():
    return render_template(
        'commandes_da.html',
        destinations=DESTINATIONS,
        dest_colors=DEST_COLORS_HEX
    )


# ══════════════════════════════════════════════════════════════
#  API STOCK / SCAN / COMMANDES CLASSIQUES
# ══════════════════════════════════════════════════════════════

@bp.route('/api/stock')
def api_stock():
    produits = Produit.query.order_by(Produit.categorie, Produit.nom).all()
    return jsonify([
        {
            'code':      p.code,
            'nom':       p.nom,
            'categorie': p.categorie,
            'stock':     p.stock,
            'stock_provisoire': calculer_stock_provisoire(p.code),
        }
        for p in produits
    ])


@bp.route('/api/scan', methods=['POST'])
def api_scan():
    data = request.json
    code = normalise_code(data.get('code', ''))

    produit = Produit.query.filter_by(code=code).first()
    if not produit:
        return jsonify({'found': False})

    return jsonify({
        'found':     True,
        'code':      produit.code,
        'nom':       produit.nom,
        'categorie': produit.categorie,
        'stock':     produit.stock,
    })


@bp.route('/api/valider', methods=['POST'])
def api_valider():
    data        = request.json
    destination = data.get('destination')
    produits    = data.get('produits', [])

    if not destination or destination not in DESTINATIONS:
        return jsonify({'ok': False, 'error': 'Destination invalide'})
    if not produits:
        return jsonify({'ok': False, 'error': 'Aucun produit'})

    now    = datetime.now()
    cmd_id = nouvelle_commande_id()
    alertes = []

    for p in produits:
        code = normalise_code(p.get('code', ''))
        qte  = p.get('quantite', 0)
        if qte <= 0:
            continue

        produit = Produit.query.filter_by(code=code).first()
        if not produit:
            continue

        produit.stock = max(0, produit.stock - qte)
        if produit.stock <= 0:
            alertes.append(f"{produit.nom} ({code})")

        db.session.add(Historique(
            cmd_id      = cmd_id,
            date        = now.strftime('%Y-%m-%d'),
            heure       = now.strftime('%H:%M:%S'),
            destination = destination,
            code        = code,
            nom         = produit.nom,
            quantite    = qte,
            stock_apres = produit.stock,
        ))

    db.session.commit()

    return jsonify({
        'ok':      True,
        'cmd_id':  cmd_id,
        'alertes': alertes,
    })


@bp.route('/api/historique')
def api_historique():
    return jsonify(load_historique())


@bp.route('/api/journalier')
def api_journalier():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    rows = Historique.query.filter_by(date=date)\
                           .order_by(Historique.heure.desc()).all()

    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':          r.cmd_id,
                'date':        r.date,
                'heure':       r.heure,
                'destination': r.destination,
                'produits':    [],
                'nb_articles': 0,
            }
        commandes[r.cmd_id]['produits'].append({
            'code':        r.code,
            'nom':         r.nom,
            'quantite':    r.quantite,
            'stock_apres': r.stock_apres,
        })
        commandes[r.cmd_id]['nb_articles'] += r.quantite

    return jsonify({
        'date':      date,
        'commandes': sorted(
            commandes.values(),
            key=lambda c: (c['date'], c['heure']),
            reverse=True
        ),
    })


# ══════════════════════════════════════════════════════════════
#  PDF
# ══════════════════════════════════════════════════════════════

@bp.route('/api/pdf/commande/<cmd_id>')
def api_pdf_commande(cmd_id):
    rows = Historique.query.filter_by(cmd_id=cmd_id).all()
    if not rows:
        return 'Commande introuvable', 404

    commande = {
        'id':          cmd_id,
        'date':        rows[0].date,
        'heure':       rows[0].heure,
        'destination': rows[0].destination,
        'produits': [
            {
                'code':        r.code,
                'nom':         r.nom,
                'quantite':    r.quantite,
                'stock_apres': r.stock_apres,
            }
            for r in rows
        ],
    }

    buf = generate_commande_pdf(commande)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'{cmd_id}.pdf')


@bp.route('/api/pdf/journalier')
def api_pdf_journalier():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    rows = Historique.query.filter_by(date=date)\
                           .order_by(Historique.destination, Historique.heure).all()

    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':          r.cmd_id,
                'date':        r.date,
                'heure':       r.heure,
                'destination': r.destination,
                'produits':    [],
            }
        commandes[r.cmd_id]['produits'].append({
            'code':        r.code,
            'nom':         r.nom,
            'quantite':    r.quantite,
            'stock_apres': r.stock_apres,
        })

    buf = generate_journalier_pdf(date, list(commandes.values()))
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'journalier_{date}.pdf')


# ══════════════════════════════════════════════════════════════
#  API COMMANDES DA
# ══════════════════════════════════════════════════════════════

@bp.route('/api/commandes_da/nouvelle', methods=['POST'])
def api_nouvelle_commande_da():
    try:
        data     = request.json
        dest     = data.get('destination')
        produits = data.get('produits', [])

        if not dest or dest not in DESTINATIONS:
            return jsonify({'ok': False, 'error': 'Destination invalide'})
        if not produits:
            return jsonify({'ok': False, 'error': 'Aucun produit'})

        now    = datetime.now()
        cmd_id = 'DA-' + now.strftime('%Y%m%d-%H%M%S')

        date_retrait = data.get('date_retrait')
        if date_retrait:
            try:
                date_retrait = datetime.strptime(date_retrait, '%Y-%m-%d').date()
            except:
                date_retrait = None
        else:
            date_retrait = None

        for p in produits:
            if p.get('quantite', 0) > 0:
                db.session.add(CommandeDA(
                    cmd_id       = cmd_id,
                    date         = now.strftime('%Y-%m-%d'),
                    heure        = now.strftime('%H:%M:%S'),
                    destination  = dest,
                    statut       = 'en_attente',
                    code         = p['code'],
                    nom          = p['nom'],
                    quantite     = p['quantite'],
                    date_retrait = date_retrait
                ))

        db.session.commit()

        commande_mail = {
            'id':          cmd_id,
            'date':        now.strftime('%Y-%m-%d'),
            'heure':       now.strftime('%H:%M'),
            'destination': dest,
            'produits':    [p for p in produits if p.get('quantite', 0) > 0],
        }

        response = jsonify({'ok': True, 'cmd_id': cmd_id})

        import threading
        threading.Thread(
            target=envoyer_mail_commande,
            args=(commande_mail, 'nouvelle'),
            daemon=True
        ).start()

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/commandes_da/modifier', methods=['POST'])
def api_modifier_commande_da():
    data     = request.json
    cmd_id   = data.get('cmd_id')
    produits = data.get('produits', [])

    if not cmd_id:
        return jsonify({'ok': False, 'error': 'ID commande manquant'})

    existing = CommandeDA.query.filter_by(cmd_id=cmd_id).first()
    if not existing:
        return jsonify({'ok': False, 'error': 'Commande introuvable'})
    dest = existing.destination

    CommandeDA.query.filter_by(cmd_id=cmd_id).delete()

    now = datetime.now()
    for p in produits:
        if p.get('quantite', 0) > 0:
            db.session.add(CommandeDA(
                cmd_id      = cmd_id,
                date        = now.strftime('%Y-%m-%d'),
                heure       = now.strftime('%H:%M:%S'),
                destination = dest,
                statut      = 'en_attente',
                code        = p['code'],
                nom         = p['nom'],
                quantite    = p['quantite'],
            ))

    db.session.commit()

    commande_mail = {
        'id':          cmd_id,
        'date':        now.strftime('%Y-%m-%d'),
        'heure':       now.strftime('%H:%M'),
        'destination': dest,
        'produits':    [p for p in produits if p.get('quantite', 0) > 0],
    }

    # ✅ On répond d'abord IMMEDIATEMENT à l'utilisateur
    response = jsonify({'ok': True})

    # ✅ Puis on envoie le mail EN ARRIERE PLAN sans attendre
    import threading
    threading.Thread(
        target=envoyer_mail_commande,
        args=(commande_mail, 'modifiée'),
        daemon=True
    ).start()

    return response


@bp.route('/api/commandes_da/supprimer', methods=['POST'])
def api_supprimer_commande_da():
    data   = request.json
    cmd_id = data.get('cmd_id')

    if not cmd_id:
        return jsonify({'ok': False, 'error': 'ID commande manquant'})

    deleted = CommandeDA.query.filter_by(cmd_id=cmd_id).delete()
    db.session.commit()

    if deleted == 0:
        return jsonify({'ok': False, 'error': 'Commande introuvable'})
    return jsonify({'ok': True})


@bp.route('/api/commandes_da')
def api_commandes_da():
    rows = CommandeDA.query.order_by(
        CommandeDA.date.desc(), CommandeDA.heure.desc()
    ).all()
    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':          r.cmd_id,
                'date':        r.date,
                'heure':       r.heure,
                'destination': r.destination,
                'statut':      r.statut,
                'produits':    [],
            }
        commandes[r.cmd_id]['produits'].append({
            'code':     r.code,
            'nom':      r.nom,
            'quantite': r.quantite,
        })
    return jsonify({'ok': True, 'commandes': list(commandes.values())})


@bp.route('/api/commandes_da/pdf/<cmd_id>')
def api_pdf_commande_da(cmd_id):
    rows = CommandeDA.query.filter_by(cmd_id=cmd_id).all()
    if not rows:
        return 'Commande introuvable', 404

    commande = {
        'id':          cmd_id,
        'date':        rows[0].date,
        'heure':       rows[0].heure,
        'destination': rows[0].destination,
        'produits': [
            {'code': r.code, 'nom': r.nom, 'quantite': r.quantite}
            for r in rows
        ],
    }

    buf = generate_commande_da_pdf(commande)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'{cmd_id}.pdf')


@bp.route('/api/catalogue')
def api_catalogue():
    produits = Produit.query.order_by(Produit.categorie, Produit.nom).all()
    return jsonify({
        'ok': True,
        'produits': [
            {
                'code': p.code,
                'nom': p.nom,
                'categorie': p.categorie,
                'stock': p.stock,
                'stock_provisoire': calculer_stock_provisoire(p.code),
            }
            for p in produits
        ]
    })


@bp.route('/stock_info')
def stock_info():
    total    = Produit.query.count()
    ruptures = Produit.query.filter(Produit.stock <= 0).count()
    return jsonify({
        'ok':             True,
        'total_produits': total,
        'ruptures':       ruptures
    })


# ══════════════════════════════════════════════════════════════
#  AJOUT / MODIFICATION PRODUIT
# ══════════════════════════════════════════════════════════════

@bp.route('/api/categories')
def api_categories():
    """Retourne la liste des catégories distinctes."""
    categories = db.session.query(Produit.categorie)\
        .distinct()\
        .filter(Produit.categorie != '')\
        .order_by(Produit.categorie).all()
    return jsonify({
        'ok': True,
        'categories': [c[0] for c in categories]
    })


@bp.route('/api/produit/nouveau', methods=['POST'])
def api_nouveau_produit():
    try:
        data = request.json
        code = data.get('code', '').strip().upper()
        nom  = data.get('nom', '').strip()
        cat  = data.get('categorie', '').strip()
        stock_init = int(data.get('stock', 0))

        if not code or not nom:
            return jsonify({'ok': False, 'error': 'Code et nom requis'}), 400

        # Vérifier si le code existe déjà
        existing = Produit.query.get(code)
        if existing:
            return jsonify({'ok': False, 'error': f'Le code "{code}" existe déjà'}), 409

        produit = Produit(
            code=code,
            nom=nom,
            categorie=cat or '',
            stock=stock_init,
            stock_mini=0,
            stock_maxi=0,
        )
        db.session.add(produit)
        db.session.commit()

        return jsonify({'ok': True, 'produit': produit.to_dict()})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  ANCIENNES ROUTES (compatibilité)
# ══════════════════════════════════════════════════════════════

@bp.route('/scanner', methods=['POST'])
def scanner():
    data        = request.json
    destination = data.get('destination')
    code        = normalise_code(data.get('code', ''))

    if not destination or destination not in DESTINATIONS:
        return jsonify({'ok': False, 'error': 'Destination invalide'})

    produit = Produit.query.get(code)
    if not produit:
        return jsonify({'ok': False, 'error': f'Code {code} introuvable'})

    return jsonify({
        'ok':      True,
        'code':    code,
        'nom':     produit.nom,
        'stock':   produit.stock,
        'warning': produit.stock <= 0
    })


@bp.route('/valider_commande', methods=['POST'])
def valider_commande():
    data        = request.json
    destination = data.get('destination')
    produits    = data.get('produits', [])

    if not destination or destination not in DESTINATIONS:
        return jsonify({'ok': False, 'error': 'Destination invalide'})
    if not produits:
        return jsonify({'ok': False, 'error': 'Aucun produit'})

    now    = datetime.now()
    cmd_id = nouvelle_commande_id()

    for p in produits:
        code    = normalise_code(p.get('code', ''))
        qte     = int(p.get('quantite', 0))
        produit = Produit.query.get(code)
        if not produit:
            continue
        produit.stock -= qte
        stock_apres    = produit.stock

        db.session.add(Historique(
            cmd_id      = cmd_id,
            date        = now.strftime('%Y-%m-%d'),
            heure       = now.strftime('%H:%M:%S'),
            destination = destination,
            code        = code,
            nom         = produit.nom,
            quantite    = qte,
            stock_apres = stock_apres,
        ))

    db.session.commit()
    return jsonify({'ok': True, 'cmd_id': cmd_id})


@bp.route('/export_pdf/<cmd_id>')
def export_pdf(cmd_id):
    commandes = load_historique()
    commande  = next((c for c in commandes if c['id'] == cmd_id), None)
    if not commande:
        return "Commande introuvable", 404
    buffer   = generate_commande_pdf(commande)
    filename = f"bon_{cmd_id}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename, mimetype='application/pdf')


@bp.route('/export_journalier')
def export_journalier():
    date_str       = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    toutes         = load_historique()
    commandes_jour = [c for c in toutes if c['date'] == date_str]
    if not commandes_jour:
        return "Aucune commande ce jour", 404
    buffer   = generate_journalier_pdf(date_str, commandes_jour)
    filename = f"journalier_{date_str}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename, mimetype='application/pdf')

@bp.route('/commandes_da_liste')
def commandes_da_liste():
    lignes = CommandeDA.query.filter_by(statut='en_attente').all()

    # Grouper par cmd_id
    grouped = {}
    for li in lignes:
        if li.cmd_id not in grouped:
            grouped[li.cmd_id] = {
                'id':          li.cmd_id,
                'destination': li.destination,
                'date':        li.date,
                'heure':       li.heure,
                'date_retrait': li.date_retrait,
                'produits':    []
            }
        p = Produit.query.get(li.code)
        grouped[li.cmd_id]['produits'].append({
            'code':     li.code,
            'nom':      li.nom,
            'quantite': li.quantite,
            'stock':    p.stock if p else 0
        })

    result = sorted(grouped.values(), key=lambda x: x['date'] + x['heure'], reverse=True)
    return jsonify({'commandes': result})


@bp.route('/valider_commande_da', methods=['POST'])
def valider_commande_da():
    cmd_id = request.json.get('cmd_id')
    lignes = CommandeDA.query.filter_by(cmd_id=cmd_id).all()
    if not lignes:
        return jsonify({'ok': False, 'error': 'Commande introuvable'})

    for li in lignes:
        li.statut = 'validee'
    db.session.commit()

    # Envoyer notification email
    now = datetime.now()
    commande_mail = {
        'id':          cmd_id,
        'date':        lignes[0].date,
        'heure':       lignes[0].heure,
        'destination': lignes[0].destination,
        'produits':    [
            {'code': li.code, 'nom': li.nom, 'quantite': li.quantite}
            for li in lignes
        ],
    }
    import threading
    threading.Thread(
        target=envoyer_mail_commande,
        args=(commande_mail, 'nouvelle'),
        daemon=True
    ).start()

    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════
#  API COMMANDES FOURNISSEUR
# ══════════════════════════════════════════════════════════════

@bp.route('/commande_fournisseur')
def commande_fournisseur_page():
    return render_template('commande_fournisseur.html')


@bp.route('/api/suggestions_reappro')
def api_suggestions_reappro():
    """Retourne les produits dont le stock est inférieur au stock minimum."""
    produits = Produit.query.filter(
        Produit.stock_mini > 0,
        Produit.stock < Produit.stock_mini
    ).order_by(Produit.categorie, Produit.nom).all()
    return jsonify({
        'ok': True,
        'suggestions': [
            {
                'code': p.code,
                'nom': p.nom,
                'categorie': p.categorie,
                'stock': p.stock,
                'stock_mini': p.stock_mini,
                'stock_maxi': p.stock_maxi,
            }
            for p in produits
        ]
    })


@bp.route('/api/commande_fournisseur/nouvelle', methods=['POST'])
def api_nouvelle_commande_fournisseur():
    try:
        data     = request.json
        produits = data.get('produits', [])

        if not produits:
            return jsonify({'ok': False, 'error': 'Aucun produit'})

        now    = datetime.now()
        cmd_id = 'FOUR-' + now.strftime('%Y%m%d-%H%M%S')

        for p in produits:
            qte = p.get('quantite', 0)
            if qte > 0:
                db.session.add(CommandeFournisseur(
                    cmd_id   = cmd_id,
                    date     = now.strftime('%Y-%m-%d'),
                    heure    = now.strftime('%H:%M:%S'),
                    code     = p['code'],
                    nom      = p['nom'],
                    quantite = qte,
                ))

        db.session.commit()

        return jsonify({'ok': True, 'cmd_id': cmd_id})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/commande_fournisseur/reception', methods=['POST'])
def api_reception_commande_fournisseur():
    """Valide la réception d'une commande fournisseur et met à jour le stock."""
    try:
        data = request.json
        cmd_id = data.get('cmd_id')
        produits = data.get('produits', [])

        if not cmd_id or not produits:
            return jsonify({'ok': False, 'error': 'Données manquantes'}), 400

        lignes = CommandeFournisseur.query.filter_by(cmd_id=cmd_id).all()
        if not lignes:
            return jsonify({'ok': False, 'error': 'Commande introuvable'}), 404

        for p in produits:
            code = p.get('code')
            quantite_recue = int(p.get('quantite_recue', 0))
            rupture = p.get('rupture', False)

            # Mettre à jour chaque ligne de la commande
            for ligne in lignes:
                if ligne.code == code:
                    ligne.statut = 'recue'
                    ligne.quantite_recue = quantite_recue
                    ligne.rupture = rupture

            # Mettre à jour le stock
            if quantite_recue > 0:
                produit = Produit.query.get(code)
                if produit:
                    produit.stock += quantite_recue

        db.session.commit()
        return jsonify({'ok': True, 'message': f'Réception validée pour {cmd_id}'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/commandes_fournisseur')
def api_commandes_fournisseur():
    rows = CommandeFournisseur.query.order_by(
        CommandeFournisseur.date.desc(),
        CommandeFournisseur.heure.desc()
    ).all()

    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':       r.cmd_id,
                'date':     r.date,
                'heure':    r.heure,
                'statut':   'recue',
                'produits': [],
            }
        # Récupérer la catégorie depuis le catalogue
        produit = Produit.query.get(r.code)
        categorie = produit.categorie if produit else ''
        if r.statut != 'recue':
            commandes[r.cmd_id]['statut'] = 'en_attente'
        commandes[r.cmd_id]['produits'].append({
            'code':      r.code,
            'nom':       r.nom,
            'quantite':  r.quantite,
            'categorie': categorie,
            'statut':    r.statut,
            'quantite_recue': r.quantite_recue,
            'rupture':   r.rupture,
        })

    return jsonify({'ok': True, 'commandes': list(commandes.values())})


@bp.route('/api/commande_fournisseur/pdf/<cmd_id>')
def api_pdf_commande_fournisseur(cmd_id):
    rows = CommandeFournisseur.query.filter_by(cmd_id=cmd_id).all()
    if not rows:
        return 'Commande introuvable', 404

    commande = {
        'id':       cmd_id,
        'date':     rows[0].date,
        'heure':    rows[0].heure,
        'produits': [
            {'code': r.code, 'nom': r.nom, 'quantite': r.quantite}
            for r in rows
        ],
    }

    buf = generate_commande_fournisseur_pdf(commande)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'{cmd_id}.pdf')


@bp.route('/api/commande_fournisseur/excel/<cmd_id>')
def api_excel_commande_fournisseur(cmd_id):
    """Génère un Excel simple et imprimable, groupé par catégorie, tenant sur 1 page."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO

    rows = CommandeFournisseur.query.filter_by(cmd_id=cmd_id).all()
    if not rows:
        return 'Commande introuvable', 404

    # Récupérer les infos avec catégorie
    enriched = []
    for r in rows:
        produit = Produit.query.get(r.code)
        enriched.append({
            'code': r.code,
            'nom': r.nom,
            'categorie': produit.categorie if produit else '',
            'quantite': r.quantite,
        })

    # Trier par catégorie puis par nom
    enriched.sort(key=lambda x: (x['categorie'] or '', x['nom'] or ''))

    # Grouper par catégorie
    categories = {}
    for p in enriched:
        cat = p['categorie'] or 'DIVERS'
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commande"

    # ═══ MISE EN PAGE POUR IMPRESSION ═══
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Styles simples
    base_font = Font(name='Calibri', size=9)
    bold_font = Font(name='Calibri', size=9, bold=True)
    title_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    qty_font = Font(name='Calibri', size=9, bold=True)
    cat_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill("solid", fgColor="2C3E50")
    cat_fill = PatternFill("solid", fgColor="7F8C8D")
    total_fill = PatternFill("solid", fgColor="2C3E50")

    # Largeurs colonnes serrées pour tenir sur une page
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 8

    row = 1

    # ── Ligne 1 : Titre ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="COMMANDE FOURNISSEUR")
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Ligne 2 : Infos ──
    total_articles = sum(p['quantite'] for p in enriched)
    ws.cell(row=row, column=1, value=f"Date : {rows[0].date}").font = bold_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=f"N° {cmd_id}").font = base_font
    ws.cell(row=row, column=4, value=f"Total :").font = bold_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    ws.cell(row=row, column=5, value=f"{total_articles} art. / {len(enriched)} réf.").font = bold_font
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Ligne vide ──
    row += 1

    # ── En-têtes colonnes ──
    for col, label in [(1, 'REF'), (2, 'DESIGNATION'), (3, 'QTÉ'),
                        (4, 'REF'), (5, 'DESIGNATION'), (6, 'QTÉ')]:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(name='Calibri', size=8, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Produits groupés par catégorie en 2 colonnes ──
    for cat_name in sorted(categories.keys()):
        produits_cat = categories[cat_name]

        # Ligne catégorie
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"  {cat_name}")
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 16
        row += 1

        # Répartir en 2 colonnes
        half = (len(produits_cat) + 1) // 2
        col1 = produits_cat[:half]
        col2 = produits_cat[half:]
        max_rows_cat = max(len(col1), len(col2))

        for i in range(max_rows_cat):
            if i < len(col1):
                p = col1[i]
                ws.cell(row=row, column=1, value=p['code']).font = base_font
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2, value=p['nom']).font = base_font
                ws.cell(row=row, column=2).border = thin_border
                qty_cell = ws.cell(row=row, column=3, value=p['quantite'])
                qty_cell.font = qty_font
                qty_cell.border = thin_border
                qty_cell.alignment = Alignment(horizontal='center')

            if i < len(col2):
                p = col2[i]
                ws.cell(row=row, column=4, value=p['code']).font = base_font
                ws.cell(row=row, column=4).border = thin_border
                ws.cell(row=row, column=5, value=p['nom']).font = base_font
                ws.cell(row=row, column=5).border = thin_border
                qty_cell = ws.cell(row=row, column=6, value=p['quantite'])
                qty_cell.font = qty_font
                qty_cell.border = thin_border
                qty_cell.alignment = Alignment(horizontal='center')

            ws.row_dimensions[row].height = 14
            row += 1

    # ── Total ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="TOTAL")
    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = thin_border
    for c in [2, 3, 4]:
        ws.cell(row=row, column=c).fill = total_fill
        ws.cell(row=row, column=c).border = thin_border
    total_val = ws.cell(row=row, column=5, value=total_articles)
    total_val.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    total_val.fill = total_fill
    total_val.alignment = Alignment(horizontal='center', vertical='center')
    total_val.border = thin_border
    total_ref = ws.cell(row=row, column=6, value=f"{len(enriched)} réf")
    total_ref.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    total_ref.fill = total_fill
    total_ref.alignment = Alignment(horizontal='center', vertical='center')
    total_ref.border = thin_border
    ws.row_dimensions[row].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=f'commande_fournisseur_{cmd_id}.xlsx',
        as_attachment=True
    )


@bp.route('/api/commande_fournisseur/csv/<cmd_id>')
def api_csv_commande_fournisseur(cmd_id):
    import csv, io
    rows = CommandeFournisseur.query.filter_by(cmd_id=cmd_id).all()
    if not rows:
        return 'Commande introuvable', 404

    # Récupérer la catégorie pour chaque produit
    enriched = []
    for r in rows:
        produit = Produit.query.get(r.code)
        enriched.append({
            'code': r.code,
            'nom': r.nom,
            'categorie': produit.categorie if produit else '',
            'quantite': r.quantite,
        })

    # Trier par catégorie puis par nom
    enriched.sort(key=lambda x: (x['categorie'] or '', x['nom'] or ''))

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # En-tête de la commande
    writer.writerow(['COMMANDE FOURNISSEUR', f'#{cmd_id}'])
    writer.writerow(['Date', f"{rows[0].date} {rows[0].heure}"])
    writer.writerow([])
    writer.writerow(['Catégorie', 'Code', 'Produit', 'Quantité'])

    categorie_courante = None
    for p in enriched:
        cat = p['categorie'] or '(Sans catégorie)'
        if cat != categorie_courante:
            writer.writerow([f'── {cat} ──', '', '', ''])
            categorie_courante = cat
        writer.writerow([cat, p['code'], p['nom'], p['quantite']])

    # Ligne de total
    total_qte = sum(p['quantite'] for p in enriched)
    writer.writerow([])
    writer.writerow(['TOTAL', '', '', total_qte])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        download_name=f'{cmd_id}.csv',
        as_attachment=True
    )


@bp.route('/api/commandes_fournisseur/csv')
def api_csv_commandes_fournisseur():
    import csv, io
    rows = CommandeFournisseur.query.order_by(
        CommandeFournisseur.date.desc(),
        CommandeFournisseur.heure.desc()
    ).all()

    # Récupérer la catégorie pour chaque produit et regrouper par commande
    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id': r.cmd_id,
                'date': r.date,
                'heure': r.heure,
                'produits': [],
            }
        produit = Produit.query.get(r.code)
        commandes[r.cmd_id]['produits'].append({
            'code': r.code,
            'nom': r.nom,
            'categorie': produit.categorie if produit else '',
            'quantite': r.quantite,
        })

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # En-tête général
    writer.writerow(['HISTORIQUE COMPLET - COMMANDES FOURNISSEUR'])
    writer.writerow([])
    writer.writerow(['Commande', 'Date', 'Heure', 'Catégorie', 'Code', 'Produit', 'Quantité'])

    ordre_commandes = sorted(commandes.values(), key=lambda c: c['date'] + c['heure'], reverse=True)
    for cmd in ordre_commandes:
        # Trier les produits par catégorie
        cmd['produits'].sort(key=lambda x: (x['categorie'] or '', x['nom'] or ''))
        categorie_courante = None
        for p in cmd['produits']:
            cat = p['categorie'] or '(Sans catégorie)'
            if cat != categorie_courante:
                writer.writerow([cmd['id'], cmd['date'], cmd['heure'], f'── {cat} ──', '', '', ''])
                categorie_courante = cat
            writer.writerow([cmd['id'], cmd['date'], cmd['heure'], cat, p['code'], p['nom'], p['quantite']])
        writer.writerow([])  # ligne vide entre chaque commande

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        download_name='commandes_fournisseur.csv',
        as_attachment=True
    )


# ══════════════════════════════════════════════════════════════
#  ROUTE DYNAMIQUE EN DERNIER ⚠️
# ══════════════════════════════════════════════════════════════

@bp.route('/api/commandes_da/<destination>')
def api_commandes_da_par_dest(destination):
    rows = CommandeDA.query.filter_by(destination=destination)\
                           .order_by(CommandeDA.date.desc(), CommandeDA.heure.desc()).all()

    commandes = {}
    for r in rows:
        if r.cmd_id not in commandes:
            commandes[r.cmd_id] = {
                'id':          r.cmd_id,
                'date':        r.date,
                'heure':       r.heure,
                'destination': r.destination,
                'statut':      r.statut,
                'produits':    [],
            }
        commandes[r.cmd_id]['produits'].append({
            'code':     r.code,
            'nom':      r.nom,
            'quantite': r.quantite,
        })

    return jsonify({'ok': True, 'commandes': list(commandes.values())})

