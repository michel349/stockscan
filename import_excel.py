import openpyxl
from app import app, db
from models import Produit, Historique


def normalise_code(code):
    c = str(code).strip()
    if 'E+' in c or 'e+' in c:
        try:
            c = str(int(float(c)))
        except:
            pass
    return c


def importer():
    wb = openpyxl.load_workbook('stock.xlsx')

    # ── Import Catalogue ──
    ws = wb["Catalogue"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        code_raw = str(row[0]).strip()
        if code_raw.lower() in ['ref', 'code', 'n°', '']:
            continue
        if row[1] is None:
            continue

        code = normalise_code(code_raw)

        try:
            nom = str(row[1]).strip() if row[1] else ''
            categorie = str(row[2]).strip() if row[2] else ''
            stock = int(row[3]) if row[3] not in (None, '') else 0
            stock_mini = int(row[4]) if row[4] not in (None, '') else 0
            stock_maxi = int(row[5]) if row[5] not in (None, '') else 0
        except Exception as e:
            print(f"  ⚠ Ligne ignorée ({row[1]}) : {e}")
            continue

        # Met à jour si existe déjà, sinon crée
        produit = Produit.query.get(code)
        if produit:
            produit.nom = nom
            produit.categorie = categorie
            produit.stock = stock
            produit.stock_mini = stock_mini
            produit.stock_maxi = stock_maxi
        else:
            db.session.add(Produit(
                code=code, nom=nom, categorie=categorie,
                stock=stock, stock_mini=stock_mini, stock_maxi=stock_maxi
            ))
        count += 1

    db.session.commit()
    print(f"  ✅ {count} produits importés")

    # ── Import Historique ──
    if "Historique" in wb.sheetnames:
        ws_h = wb["Historique"]
        count_h = 0
        for row in ws_h.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cmd_id = str(row[0]).strip()
            if cmd_id.lower() in ['id commande', 'cmd_id', 'id']:
                continue

            try:
                db.session.add(Historique(
                    cmd_id=cmd_id,
                    date=str(row[1]) if row[1] else '',
                    heure=str(row[2]) if row[2] else '',
                    destination=str(row[3]) if row[3] else '',
                    code=str(row[4]) if row[4] else '',
                    nom=str(row[5]) if row[5] else '',
                    quantite=int(row[6]) if row[6] not in (None, '') else 0,
                    stock_apres=int(row[7]) if row[7] not in (None, '') else 0,
                ))
                count_h += 1
            except Exception as e:
                print(f"  ⚠ Historique ignoré ({cmd_id}) : {e}")
                continue

        db.session.commit()
        print(f"  ✅ {count_h} lignes d'historique importées")

    wb.close()
    print("\n  🎉 Import terminé !")


if __name__ == '__main__':
    with app.app_context():
        print("\n" + "=" * 50)
        print("  📥 IMPORT EXCEL → BASE DE DONNÉES")
        print("=" * 50)
        importer()
